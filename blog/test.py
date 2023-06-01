from openpyxl import Workbook
from openpyxl.styles import PatternFill

# データリストの定義
data_list = [1, 2]

# Workbookの作成
wb = Workbook()
ws = wb.active

# データの書き込み
for i, data in enumerate(data_list, start=1):
    ws.cell(row=i, column=1, value=data)

# 背景色の設定
fill = PatternFill(fill_type='solid', fgColor='FFFF00')  # 黄色の背景色を設定

# 偶数行に背景色を設定
for i in range(2, len(data_list)+1, 2):
    ws.cell(row=i, column=1).fill = fill

# 保存
wb.save('output.xlsx')
